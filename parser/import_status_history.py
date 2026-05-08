"""
Импорт помесячной истории статусов помещений (Сдан / Не сдан) из годовых
Excel-файлов план-факт со шары Бюджета.

Что делает:
  1. Идёт по «главным» файлам каждого года 2016-2026 (план-факт / заполняем
     эту форму / рабочая таблица).
  2. Для каждого xlsx ищет помесячные листы — название содержит месяц
     («январь 22», «п-ф февраль», «март 2024», «декабрь» и т.д.).
  3. На каждом листе автодетектит колонку со статусом — первая колонка,
     где встречаются слова «сдан» / «не сдан». Соседние колонки определяет
     по близости: floor → room → area → status → rate → legal → trade.
  4. Год месяца определяет: либо из имени листа (на нём явно указан год —
     'январь 23'), либо из имени файла. Это спасает от ситуации, когда
     2024-файл содержит 'январь 25', а 2026-файл — 'декабрь' (т.е. 2025).
  5. Пишет в rent_status_history с UNIQUE по (year, month, room, source_file).

Запуск:
  python import_status_history.py                 # все годы
  python import_status_history.py --year 2024     # один файл
  python import_status_history.py --dry-run
"""
from __future__ import annotations

import argparse
import logging
import os
import re
import sqlite3
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

try:
    import openpyxl
    from openpyxl.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:
    print('Нужен openpyxl: pip install openpyxl', file=sys.stderr)
    sys.exit(2)

HERE = Path(__file__).parent.resolve()
DB_PATH = Path(os.environ.get('DASHBOARD_DB', str(HERE.parent / 'dashboard.db')))
DEFAULT_ROOT = r'\\Acad-server\общие\02_Бухгалтерия\01_Бюджет'
LOG_DIR = HERE / 'logs'
LOG_DIR.mkdir(exist_ok=True)


# ── маппинг месяцев из имени листа ─────────────────────────────────────
MONTH_RU = {
    'январ': 1, 'феврал': 2, 'март': 3, 'апрел': 4, 'мая': 5, 'май': 5,
    'июн': 6, 'июл': 7, 'август': 8, 'сентябр': 9, 'октябр': 10,
    'ноябр': 11, 'декабр': 12,
}
YEAR_RE = re.compile(r'(20\d{2})')


def detect_month_in_sheet_name(name: str) -> int | None:
    n = name.lower()
    for stem, m in MONTH_RU.items():
        if stem in n:
            return m
    return None


def detect_year_in_text(text: str) -> int | None:
    """Найти год в имени листа/файла. Двухзначный (22, 23) расширяем до 20XX."""
    m = YEAR_RE.search(text)
    if m:
        return int(m.group(1))
    # «январь 22» или «п-ф январь » — ищем 2-значный год.
    m2 = re.search(r'\b(\d{2})\b', text)
    if m2:
        return 2000 + int(m2.group(1))
    return None


def detect_year(sheet_name: str, file_path: Path) -> int | None:
    """Сначала пробуем взять год из имени листа («январь 25»), потом из
    имени файла («рабочая таблица 2024.xlsx»). Пути не помогают — там
    всегда указан «свой» год папки, что не всегда совпадает с фактическим
    периодом листа."""
    return (detect_year_in_text(sheet_name)
            or detect_year_in_text(file_path.stem)
            or detect_year_in_text(str(file_path.parent.name)))


def is_relevant_sheet(name: str) -> bool:
    n = name.lower()
    if not detect_month_in_sheet_name(n):
        return False
    # Исключаем «съезд-заезд», «реестр ТО», «должники сети» и пр. — там
    # тоже могут встречаться слова «сдан», но структура другая.
    bad = ('съезд', 'заезд', 'итого', 'свод', 'годов', 'график')
    return not any(b in n for b in bad)


# ── маркеры статуса ────────────────────────────────────────────────────
def normalize_status(value: Any) -> tuple[str, str] | None:
    """Возвращает (canonical, raw) или None если значение не статус.
    canonical ∈ {'rented', 'not_rented', 'other'}"""
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    low = s.lower()
    if 'не сдан' in low or 'не сдано' in low or 'свобод' in low:
        return 'not_rented', s
    if low in ('сдан', 'сдано', 'занят', 'занято'):
        return 'rented', s
    if 'сдан' in low and len(low) < 30:
        # Похоже на 'сдан', но с уточнением — например 'сдан в субаренду'.
        return 'rented', s
    return None


# ── автодетект колонок ─────────────────────────────────────────────────
def find_status_column(ws: Worksheet, scan_rows: int = 60) -> int | None:
    """Колонка, в которой больше всего ячеек со словом 'сдан'/'не сдан'
    в первых scan_rows строках."""
    counts: dict[int, int] = {}
    for row in ws.iter_rows(min_row=1, max_row=scan_rows, values_only=False):
        for cell in row:
            if normalize_status(cell.value):
                counts[cell.column] = counts.get(cell.column, 0) + 1
    if not counts:
        return None
    best_col, best_n = max(counts.items(), key=lambda kv: kv[1])
    return best_col if best_n >= 5 else None


def to_float(v: Any) -> float | None:
    if v is None or v == '':
        return None
    if isinstance(v, (int, float)):
        return float(v) if v == v else None
    s = str(v).strip().replace(',', '.').replace('\xa0', '').replace(' ', '')
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def to_str(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


# ── разбор листа ───────────────────────────────────────────────────────
def parse_sheet(ws: Worksheet, source_file: Path) -> list[dict]:
    """Возвращает список словарей с полями rent_status_history."""
    status_col = find_status_column(ws)
    if not status_col:
        return []

    year = detect_year(ws.title, source_file)
    month = detect_month_in_sheet_name(ws.title)
    if not year or not month:
        return []

    # Колонки floor/room/area стоят в фиксированных позициях ВО ВСЕХ
    # форматах, что 2022+, что 2020. Различается только что идёт после
    # колонки area: в 2022+ сразу status, в 2020 сначала rate, потом status.
    #   col 1: floor (-2, -1, 1, 2, ..., УЛ)
    #   col 2: room  ('ПРК', 'R1-6', 'A26', ...)
    #   col 3: area  (число м²)
    #   status: автодетект, обычно col 4 (новый формат) или col 5 (2020)
    #   legal_name = status_col + 2 (между ними одна служебная колонка)
    #   trade_name = status_col + 3
    # Старая логика «area = status_col − 1» давала false-positive для 2020,
    # где status_col=5: area=4 = ставка, не площадь. Отсюда космические
    # 302 000 м² в 2020-01.
    floor_col = 1
    room_col  = 2
    area_col  = 3
    legal_col = status_col + 2
    trade_col = status_col + 3
    # Ставка лежит на col 5 в 2022+ (после статуса) и на col 4 в 2020 (перед).
    rate_col  = status_col - 1 if status_col >= 5 else status_col + 1

    rows: list[dict] = []
    seen_rooms: set[str] = set()  # дедуп внутри одного листа: первая встреча выигрывает

    for r, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
        if r > 600:  # хвост обычно пустой/итоги
            break
        if not row or len(row) < status_col:
            continue
        norm = normalize_status(row[status_col - 1])
        if not norm:
            continue
        status, status_raw = norm

        room = to_str(row[room_col - 1]) if len(row) >= room_col else None
        if not room:
            continue
        room_key = room.lower()
        if room_key in seen_rooms:
            continue
        seen_rooms.add(room_key)

        rows.append({
            'year': year, 'month': month,
            'floor': to_str(row[floor_col - 1]) if len(row) >= floor_col else None,
            'room': room,
            'area_m2': to_float(row[area_col - 1]) if len(row) >= area_col else None,
            'status': status, 'status_raw': status_raw,
            'rate_per_m2': to_float(row[rate_col - 1]) if len(row) >= rate_col else None,
            'legal_name': to_str(row[legal_col - 1]) if len(row) >= legal_col else None,
            'trade_name': to_str(row[trade_col - 1]) if len(row) >= trade_col else None,
        })
    return rows


# ── работа с файлами ───────────────────────────────────────────────────
MAIN_FILE_HINTS = (
    'план-факт', 'план факт', 'планфакт',
    'заполняем эту форму', 'заполняем эту таблицу',
    'рабочая таблица',
)


def is_main_file(path: Path) -> bool:
    name = path.name.lower()
    return any(h in name for h in MAIN_FILE_HINTS)


def discover_files(root: Path, year_filter: int | None = None) -> list[Path]:
    files: list[Path] = []
    for p in root.rglob('*.xlsx'):
        if p.name.startswith('~$'):
            continue
        if any(part.lower() in {'архив', 'thumbs.db', '__pycache__'} for part in p.parts):
            continue
        if not is_main_file(p):
            continue
        # Год возьмём из имени файла или ближайшей папки.
        text = p.stem + ' ' + p.parent.name
        y = detect_year_in_text(text)
        if year_filter is not None and y != year_filter:
            continue
        files.append(p)
    files.sort()
    return files


# ── работа с БД ────────────────────────────────────────────────────────
SCHEMA_SQL = '''
CREATE TABLE IF NOT EXISTS rent_status_history (
  id            INTEGER PRIMARY KEY AUTOINCREMENT,
  year          INTEGER NOT NULL,
  month         INTEGER NOT NULL,
  floor         TEXT,
  room          TEXT,
  area_m2       REAL,
  status        TEXT NOT NULL,
  status_raw    TEXT,
  legal_name    TEXT,
  trade_name    TEXT,
  rate_per_m2   REAL,
  source_file   TEXT NOT NULL,
  source_sheet  TEXT NOT NULL,
  imported_at   TEXT NOT NULL,
  UNIQUE(year, month, room, source_file)
);
CREATE INDEX IF NOT EXISTS idx_rsh_year_month ON rent_status_history(year, month);
CREATE INDEX IF NOT EXISTS idx_rsh_room       ON rent_status_history(room);
CREATE INDEX IF NOT EXISTS idx_rsh_status     ON rent_status_history(status);
'''


def open_db() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.execute('PRAGMA journal_mode = WAL')
    conn.executescript(SCHEMA_SQL)
    return conn


def now_iso() -> str:
    return datetime.now().strftime('%Y-%m-%d %H:%M:%S')


def upsert(conn: sqlite3.Connection, fields: dict, source_file: str,
           source_sheet: str, dry_run: bool) -> str:
    if dry_run:
        return 'inserted'
    existing = conn.execute(
        'SELECT id FROM rent_status_history '
        'WHERE year=? AND month=? AND COALESCE(room,"")=COALESCE(?,"") AND source_file=?',
        (fields['year'], fields['month'], fields['room'], source_file),
    ).fetchone()
    now = now_iso()
    if existing:
        conn.execute(
            'UPDATE rent_status_history SET '
            'floor=?, area_m2=?, status=?, status_raw=?, legal_name=?, trade_name=?, '
            'rate_per_m2=?, source_sheet=?, imported_at=? WHERE id=?',
            (fields['floor'], fields['area_m2'], fields['status'], fields['status_raw'],
             fields['legal_name'], fields['trade_name'], fields['rate_per_m2'],
             source_sheet, now, existing[0]),
        )
        return 'updated'
    conn.execute(
        'INSERT INTO rent_status_history '
        '(year, month, floor, room, area_m2, status, status_raw, legal_name, '
        ' trade_name, rate_per_m2, source_file, source_sheet, imported_at) '
        'VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)',
        (fields['year'], fields['month'], fields['floor'], fields['room'],
         fields['area_m2'], fields['status'], fields['status_raw'],
         fields['legal_name'], fields['trade_name'], fields['rate_per_m2'],
         source_file, source_sheet, now),
    )
    return 'inserted'


def setup_logging() -> Path:
    log_file = LOG_DIR / f'status_{datetime.now():%Y-%m-%d_%H-%M}.log'
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s [%(levelname)s] %(message)s',
        handlers=[
            logging.FileHandler(log_file, encoding='utf-8'),
            logging.StreamHandler(sys.stdout),
        ],
        force=True,
    )
    return log_file


def dedupe_per_month(conn: sqlite3.Connection) -> int:
    """В каждом (year, month) оставляем один source_file — тот, в котором
    больше строк (= наиболее полная версия). Это убирает копии и
    промежуточные снимки одного периода. Возвращает удалённые строки."""
    deleted = 0
    pairs = conn.execute(
        'SELECT DISTINCT year, month FROM rent_status_history ORDER BY year, month'
    ).fetchall()
    for year, month in pairs:
        rows = conn.execute(
            'SELECT source_file, COUNT(*) AS n FROM rent_status_history '
            'WHERE year=? AND month=? GROUP BY source_file ORDER BY n DESC, source_file',
            (year, month),
        ).fetchall()
        if len(rows) <= 1:
            continue
        canonical = rows[0][0]
        for src, n in rows[1:]:
            d = conn.execute(
                'DELETE FROM rent_status_history WHERE year=? AND month=? AND source_file=?',
                (year, month, src),
            ).rowcount
            deleted += d
            logging.info('  [%d-%02d] оставляем %s (%d), убираем %s (%d)',
                         year, month, Path(canonical).name, rows[0][1],
                         Path(src).name, n)
    return deleted


def drop_garbage_years(conn: sqlite3.Connection, min_year: int = 2019) -> int:
    """Удаляет записи с year < min_year — это false positives от случайных
    дат в исторических колонках (в листах попадаются записи 2001, 2008 и
    т.п. из старых проектных таблиц)."""
    return conn.execute(
        'DELETE FROM rent_status_history WHERE year < ?', (min_year,)
    ).rowcount


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument('--root', default=DEFAULT_ROOT)
    ap.add_argument('--year', type=int, help='только один год по имени файла')
    ap.add_argument('--dry-run', action='store_true')
    args = ap.parse_args()

    setup_logging()
    logging.info('Импорт истории статусов. Корень: %s', args.root)

    root = Path(args.root)
    if not root.exists():
        logging.error('Корень не найден: %s', root)
        return 2

    files = discover_files(root, args.year)
    logging.info('Файлов к обработке: %d', len(files))
    if not files:
        return 0

    conn = open_db()
    totals = {'inserted': 0, 'updated': 0}
    try:
        for path in files:
            logging.info('Файл: %s', path.name)
            try:
                wb: Workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
            except Exception as e:
                logging.error('  Не открылся: %s', e)
                continue
            try:
                for ws in wb.worksheets:
                    if not is_relevant_sheet(ws.title):
                        continue
                    rows = parse_sheet(ws, path)
                    if not rows:
                        continue
                    logging.info('  лист «%s»: %d строк (%d/%d)', ws.title,
                                 len(rows), rows[0]['month'], rows[0]['year'])
                    for f in rows:
                        res = upsert(conn, f, str(path), ws.title, args.dry_run)
                        totals[res] = totals.get(res, 0) + 1
                if not args.dry_run:
                    conn.commit()
            finally:
                wb.close()
    finally:
        conn.close()

    logging.info('Готово. Записей: вставлено %d, обновлено %d.',
                 totals['inserted'], totals['updated'])

    if not args.dry_run:
        conn = open_db()
        try:
            n_garbage = drop_garbage_years(conn)
            logging.info('Удалено мусорных годов (< 2019): %d', n_garbage)
            n_dedup = dedupe_per_month(conn)
            logging.info('Дедуп по (year, month): -%d строк', n_dedup)
            conn.commit()
        finally:
            conn.close()
    return 0


if __name__ == '__main__':
    sys.exit(main())
