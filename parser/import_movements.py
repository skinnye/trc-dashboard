"""
Импорт реестра съездов/заездов арендаторов из годовых Excel-файлов
\\Acad-server\\общие\\02_Бухгалтерия\\01_Бюджет\\<год>\\

Стратегия:
  1. Рекурсивно обходим папку Бюджета.
  2. Каждый .xlsx (без lock-файлов ~$...) пробуем определить год — из пути
     или имени файла.
  3. По умолчанию в каждом xlsx обрабатываются ТОЛЬКО листы с подстрокой
     'съезд' или 'заезд' в имени — это устойчивое соглашение, менеджеры
     заполняют именно такой лист (обычно «4 съезд-заезд»). На таких листах
     ищем секционные заголовки:
       'РАСТОРЖЕНИЕ ДОГОВОРА' / 'СЪЕЗД'   → kind='departure'
       'НОВЫЕ АРЕНДАТОРЫ'    / 'ЗАЕЗД'   → kind='arrival'
     После маркера читаем строки до следующего разделителя/ИТОГО/пустоты.
  4. Флаг --scan-all-sheets разрешает поиск в листах с другими названиями
     (медленнее и шумнее, использовать только для разовой ревизии).
  5. Колонки определяем из ближайшей шапки выше секции, либо из дефолта.
  6. Пишем в tenant_movements c UPSERT по UNIQUE(year, kind, seq_no, room,
     source_file). Повторный импорт ничего не плодит — обновляются поля.

Запуск:
  python import_movements.py                 # вся шара 2016-2026
  python import_movements.py --year 2022     # только один год
  python import_movements.py --root PATH     # альтернативный корень
  python import_movements.py --dry-run       # ничего не пишет, только лог
"""
from __future__ import annotations

import argparse
import logging
import os
import re
import sqlite3
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

try:
    import openpyxl
    from openpyxl.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:
    print('Нужен openpyxl: pip install openpyxl', file=sys.stderr)
    sys.exit(2)


HERE = Path(__file__).parent.resolve()
DB_PATH = Path(os.environ.get('DASHBOARD_DB', str(HERE.parent / 'dashboard.db')))
DEFAULT_ROOT = r'\\Acad-server\общие\02_Бухгалтерия\01_Бюджет'
LOG_DIR = HERE / 'logs'
LOG_DIR.mkdir(exist_ok=True)


# ── маркеры секций в листах ────────────────────────────────────────────
DEPARTURE_MARKERS = (
    'РАСТОРЖЕНИЕ ДОГОВОРА',
    'СЪЕЗД',
    'СЪЕЗДЫ',
)
ARRIVAL_MARKERS = (
    'НОВЫЕ АРЕНДАТОРЫ',
    'ЗАЕЗД',
    'ЗАЕЗДЫ',
)
END_MARKERS = (
    'ИТОГО',
    'ВСЕГО',
)

# Колонка в листе → роль. Подобрано по эталонной форме «4 съезд-заезд» 2022:
#   1=№, 2=этаж, 3=помещение, 4=площадь, 5=ставка, 6=юр.лицо, 7=бренд,
#   8=без НДС, 9=с НДС, 10=дата
# Если в файле другой порядок — берётся из шапки листа.
DEFAULT_COLS = {
    'seq_no':      1,
    'floor':       2,
    'room':        3,
    'area_m2':     4,
    'rate':        5,
    'legal_name':  6,
    'trade_name':  7,
    'charges_no_vat':   8,
    'charges_with_vat': 9,
    'event_date':  10,
}

# Ключевые слова, по которым угадываем колонки если шапка нашлась в листе.
HEADER_HINTS = {
    'seq_no':           ['№', 'no'],
    'floor':            ['этаж'],
    'room':             ['помещен', 'объект', 'кол-во дверей', 'дверей'],
    'area_m2':          ['площад', 'м2', 'м²'],
    'rate':             ['ставк', 'ср. ар'],
    'legal_name':       ['наименование арендатор', 'арендатор', 'статьи', 'юр'],
    'trade_name':       ['торгов', 'бренд', 'юр.наим'],
    'charges_no_vat':   ['без ндс'],
    'charges_with_vat': ['с ндс'],
    'event_date':       ['дата', 'апп', 'авр', 'срок'],
}


def setup_logging() -> Path:
    log_file = LOG_DIR / f'movements_{datetime.now():%Y-%m-%d_%H-%M}.log'
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s [%(levelname)s] %(message)s',
        handlers=[
            logging.FileHandler(log_file, encoding='utf-8'),
            logging.StreamHandler(sys.stdout),
        ],
        force=True,
    )
    return log_file


def now_iso() -> str:
    return datetime.now().strftime('%Y-%m-%d %H:%M:%S')


# ── разбор клеток ──────────────────────────────────────────────────────
def cell_str(v: Any) -> str:
    if v is None:
        return ''
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    return str(v).strip()


def cell_float(v: Any) -> float | None:
    if v is None or v == '':
        return None
    if isinstance(v, (int, float)):
        return float(v) if v == v else None
    s = cell_str(v).replace(',', '.').replace(' ', '').replace('\xa0', '')
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def cell_int(v: Any) -> int | None:
    f = cell_float(v)
    if f is None:
        return None
    if isinstance(v, bool):  # bool — подтип int, защищаемся
        return None
    return int(f)


def cell_date(v: Any) -> tuple[str | None, str | None]:
    """Возвращает (iso_date, raw_string).
    Если значение — datetime, парсим. Иначе возвращаем сырой текст."""
    if v is None or v == '':
        return None, None
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d'), None
    s = cell_str(v)
    if not s:
        return None, None
    # Попытка распарсить простые форматы.
    for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%d-%m-%Y'):
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d'), None
        except ValueError:
            continue
    return None, s


# ── определение года и поиск файлов ────────────────────────────────────
YEAR_RE = re.compile(r'(20\d{2})')


def detect_year(path: Path) -> int | None:
    """Год из пути или имени файла. Берём первый 20XX, который найдём."""
    for part in [path.stem, *map(str, path.parents)]:
        m = YEAR_RE.search(part)
        if m:
            return int(m.group(1))
    return None


# Имена «главных» годовых файлов — ищем по этим маркерам, чтобы не лезть
# в случайные Книги1.xlsx и помесячные отчёты.
MAIN_FILE_HINTS = (
    'план-факт', 'план факт', 'планфакт',
    'заполняем эту форму', 'заполняем эту таблицу',
    'рабочая таблица',
)


def is_main_file(path: Path) -> bool:
    name = path.name.lower()
    return any(h in name for h in MAIN_FILE_HINTS)


def discover_files(root: Path, year_filter: int | None = None,
                   scan_all: bool = False) -> list[tuple[Path, int]]:
    """Список (file_path, year). По умолчанию только «главные» годовые
    файлы (план-факт / заполняем эту форму / рабочая таблица). С --scan-all
    берёт все .xlsx (медленнее, больше шума)."""
    files: list[tuple[Path, int]] = []
    for p in root.rglob('*.xlsx'):
        name = p.name
        if name.startswith('~$'):  # Excel lock
            continue
        if any(part.lower() in {'.git', 'thumbs.db', '__pycache__', 'архив'}
               for part in p.parts):
            continue
        if not scan_all and not is_main_file(p):
            continue
        year = detect_year(p)
        if year is None:
            continue
        if year_filter is not None and year != year_filter:
            continue
        files.append((p, year))
    files.sort(key=lambda x: (x[1], str(x[0])))
    return files


# ── обнаружение секций в листе ─────────────────────────────────────────
def find_marker(text: str, markers: tuple[str, ...]) -> bool:
    upper = text.upper()
    return any(m.upper() in upper for m in markers)


def detect_columns_from_header(ws: Worksheet, header_row: int) -> dict[str, int]:
    """Если выше секции есть строка-шапка с подписями колонок — пытаемся
    разобрать её. Возвращает мапу {role: col_idx}, частичную."""
    cols: dict[str, int] = {}
    if header_row < 1:
        return cols
    row = ws[header_row]
    for cell in row:
        text = cell_str(cell.value).lower()
        if not text:
            continue
        for role, hints in HEADER_HINTS.items():
            if role in cols:
                continue
            if any(h in text for h in hints):
                cols[role] = cell.column
                break
    return cols


def looks_like_data_row(row_values: list[Any], cols: dict[str, int]) -> bool:
    """Эвристика: есть ли в строке хотя бы одно «полезное» поле — название
    арендатора, помещение или дата. Иначе строку считаем разделителем/итого."""
    keys = ('legal_name', 'room', 'event_date')
    for k in keys:
        idx = cols.get(k)
        if idx is None:
            continue
        v = row_values[idx - 1] if idx <= len(row_values) else None
        s = cell_str(v)
        if s:
            return True
    return False


def is_total_row(row_values: list[Any]) -> bool:
    text = ' '.join(cell_str(v) for v in row_values).upper()
    return any(m in text for m in END_MARKERS)


def parse_section(ws: Worksheet, start_row: int, cols: dict[str, int]) -> Iterable[tuple[int, dict]]:
    """Читает строки данных, начиная с start_row, до итога/пустого хвоста.
    Возвращает (row_index, fields)."""
    blank_streak = 0
    for r in range(start_row, ws.max_row + 1):
        row = [c.value for c in ws[r]]
        if is_total_row(row):
            return
        if not any(v not in (None, '') for v in row):
            blank_streak += 1
            if blank_streak >= 2:  # две пустые подряд = конец секции
                return
            continue
        blank_streak = 0
        if not looks_like_data_row(row, cols):
            continue

        def at(role: str):
            idx = cols.get(role)
            if idx is None or idx > len(row):
                return None
            return row[idx - 1]

        iso, raw = cell_date(at('event_date'))
        fields = {
            'seq_no':           cell_int(at('seq_no')),
            'floor':            cell_str(at('floor')) or None,
            'room':             cell_str(at('room')) or None,
            'area_m2':          cell_float(at('area_m2')),
            'rate_per_m2':      cell_float(at('rate')),
            'legal_name':       cell_str(at('legal_name')) or None,
            'trade_name':       cell_str(at('trade_name')) or None,
            'charges_no_vat':   cell_float(at('charges_no_vat')),
            'charges_with_vat': cell_float(at('charges_with_vat')),
            'event_date':       iso,
            'date_raw':         raw,
        }
        yield r, fields


def is_section_header_row(row_values: list[Any], markers: tuple[str, ...]) -> bool:
    """Считаем строку заголовком секции, только если МАРКЕР занимает
    практически всю строку — типичный паттерн merged-ячейки в реестре.
    Это спасает от false positives, когда слово 'съезд' встречается в
    комментарии длинной строки данных.

    Критерии:
      • в строке хотя бы одна непустая ячейка содержит маркер
      • эта ячейка состоит из одного «крупного» текста, длиной до 80 символов
      • суммарный текст всей строки не сильно длиннее ячейки с маркером
        (не должно быть кучи других данных в той же строке)
    """
    cells_with_text = [(i, cell_str(v)) for i, v in enumerate(row_values)
                       if cell_str(v)]
    if not cells_with_text:
        return False
    full_text = ' '.join(s for _, s in cells_with_text).upper()
    has_marker = any(m.upper() in full_text for m in markers)
    if not has_marker:
        return False
    # Проверяем "целостность" — в строке не должно быть много разных значений.
    # Из-за merge-cell openpyxl возвращает значение только в первой ячейке
    # объединения, остальные пусты — поэтому unique-текстов мало.
    unique_texts = {s for _, s in cells_with_text if s}
    if len(unique_texts) > 3:
        return False
    # Длина каждой непустой ячейки разумная (это заголовок, не данные).
    if any(len(s) > 120 for _, s in cells_with_text):
        return False
    return True


def find_sections(ws: Worksheet) -> list[tuple[str, int, dict[str, int]]]:
    """Ищет по листу секции «съезд» и «заезд». Возвращает список
    (kind, first_data_row, columns_map). Использует строгое определение
    заголовка секции — иначе на больших листах ловит false positives."""
    sections: list[tuple[str, int, dict[str, int]]] = []
    last_header_row = -1

    # Не сканируем гигантские листы целиком — для секционных реестров
    # типичные размеры до 200 строк. Если лист больше — он точно не реестр.
    scan_until = min(ws.max_row, 500)

    for r in range(1, scan_until + 1):
        row = [c.value for c in ws[r]]
        if not any(v not in (None, '') for v in row):
            continue

        is_dep = is_section_header_row(row, DEPARTURE_MARKERS)
        is_arr = is_section_header_row(row, ARRIVAL_MARKERS)
        if not is_dep and not is_arr:
            text_lower = ' '.join(cell_str(v) for v in row).lower()
            if any(k in text_lower for k in ('арендатор', 'этаж', 'помещен', 'площад', 'ставк', 'дата')):
                last_header_row = r
            continue

        if last_header_row > 0 and last_header_row < r:
            detected = detect_columns_from_header(ws, last_header_row)
        else:
            detected = {}
        cols = {**DEFAULT_COLS, **detected}
        kind = 'departure' if is_dep else 'arrival'
        sections.append((kind, r + 1, cols))

    return sections


# ── работа с БД ────────────────────────────────────────────────────────
SCHEMA_SQL = '''
CREATE TABLE IF NOT EXISTS tenant_movements (
  id              INTEGER PRIMARY KEY AUTOINCREMENT,
  year            INTEGER NOT NULL,
  kind            TEXT NOT NULL,
  seq_no          INTEGER,
  floor           TEXT,
  room            TEXT,
  area_m2         REAL,
  rate_per_m2     REAL,
  legal_name      TEXT,
  trade_name      TEXT,
  charges_no_vat  REAL,
  charges_with_vat REAL,
  event_date      TEXT,
  date_raw        TEXT,
  source_file     TEXT NOT NULL,
  source_sheet    TEXT NOT NULL,
  source_row      INTEGER,
  imported_at     TEXT NOT NULL,
  UNIQUE(year, kind, seq_no, room, source_file)
);
CREATE INDEX IF NOT EXISTS idx_tm_year_kind ON tenant_movements(year, kind);
CREATE INDEX IF NOT EXISTS idx_tm_room      ON tenant_movements(room);
CREATE INDEX IF NOT EXISTS idx_tm_date      ON tenant_movements(event_date);
'''


def open_db() -> sqlite3.Connection:
    if not DB_PATH.exists():
        raise FileNotFoundError(f'БД не найдена: {DB_PATH}')
    conn = sqlite3.connect(DB_PATH)
    conn.execute('PRAGMA journal_mode = WAL')
    # Создаём таблицу, если её ещё нет (Next.js делает то же самое при
    # первом обращении). Полностью эквивалентно по полям и индексам схеме
    # из web/src/lib/db.ts — обе стороны должны видеть одинаковый набор.
    conn.executescript(SCHEMA_SQL)
    return conn


def upsert_movement(conn: sqlite3.Connection, year: int, kind: str,
                    fields: dict, source_file: str, source_sheet: str,
                    source_row: int, dry_run: bool = False) -> str:
    """Возвращает 'inserted' / 'updated' / 'skipped'."""
    seq_no = fields.get('seq_no')
    room = fields.get('room')
    legal = fields.get('legal_name')
    if not legal and not room:
        return 'skipped'  # совсем пустая запись

    if dry_run:
        return 'inserted'  # для подсчёта

    existing = conn.execute(
        'SELECT id FROM tenant_movements '
        'WHERE year = ? AND kind = ? AND COALESCE(seq_no, -1) = COALESCE(?, -1) '
        '  AND COALESCE(room, "") = COALESCE(?, "") AND source_file = ?',
        (year, kind, seq_no, room, source_file),
    ).fetchone()
    now = now_iso()

    if existing:
        conn.execute(
            'UPDATE tenant_movements SET '
            'floor = ?, area_m2 = ?, rate_per_m2 = ?, legal_name = ?, trade_name = ?, '
            'charges_no_vat = ?, charges_with_vat = ?, event_date = ?, date_raw = ?, '
            'source_sheet = ?, source_row = ?, imported_at = ? '
            'WHERE id = ?',
            (
                fields.get('floor'), fields.get('area_m2'), fields.get('rate_per_m2'),
                fields.get('legal_name'), fields.get('trade_name'),
                fields.get('charges_no_vat'), fields.get('charges_with_vat'),
                fields.get('event_date'), fields.get('date_raw'),
                source_sheet, source_row, now, existing[0],
            ),
        )
        return 'updated'

    conn.execute(
        'INSERT INTO tenant_movements '
        '(year, kind, seq_no, floor, room, area_m2, rate_per_m2, legal_name, trade_name, '
        ' charges_no_vat, charges_with_vat, event_date, date_raw, '
        ' source_file, source_sheet, source_row, imported_at) '
        'VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
        (
            year, kind, seq_no, fields.get('floor'), room, fields.get('area_m2'),
            fields.get('rate_per_m2'), fields.get('legal_name'), fields.get('trade_name'),
            fields.get('charges_no_vat'), fields.get('charges_with_vat'),
            fields.get('event_date'), fields.get('date_raw'),
            source_file, source_sheet, source_row, now,
        ),
    )
    return 'inserted'


# ── оркестрация ────────────────────────────────────────────────────────
def sheet_name_looks_relevant(name: str) -> bool:
    """Имя листа намекает на реестр движений арендаторов."""
    n = name.lower()
    return ('съезд' in n) or ('заезд' in n)


def process_file(conn: sqlite3.Connection, path: Path, year: int,
                 dry_run: bool, scan_all_sheets: bool) -> dict[str, int]:
    """Открывает xlsx, находит секции в каждом листе, импортирует строки."""
    counters = {'inserted': 0, 'updated': 0, 'skipped': 0, 'sheets_with_data': 0}
    try:
        wb: Workbook = openpyxl.load_workbook(path, data_only=True, read_only=False)
    except Exception as e:
        logging.error('  Не удалось открыть %s: %s', path.name, e)
        return counters

    try:
        for ws in wb.worksheets:
            # По умолчанию заходим только в листы с явным именем «съезд/заезд».
            # На основной форме они всегда так и называются. На больших
            # листах (по 2-3 тыс. строк) маркеры дают false positives.
            if not scan_all_sheets and not sheet_name_looks_relevant(ws.title):
                continue
            sections = find_sections(ws)
            if not sections:
                continue
            counters['sheets_with_data'] += 1
            logging.info('  лист «%s»: %d секций', ws.title, len(sections))
            for kind, start_row, cols in sections:
                rows_in_section = 0
                for r, fields in parse_section(ws, start_row, cols):
                    res = upsert_movement(conn, year, kind, fields,
                                          str(path), ws.title, r, dry_run)
                    counters[res] = counters.get(res, 0) + 1
                    rows_in_section += 1
                logging.info('    %s: %d строк', kind, rows_in_section)
    finally:
        wb.close()
    return counters


def dedupe_per_year(conn: sqlite3.Connection) -> tuple[int, list[str]]:
    """В каждом году оставляем только источник с максимальным числом записей —
    он считается «канонической» версией файла. Это убирает копии вида
    «Копия ...», «план YYYY/...», «YYYY/05/...» где данные неполные.

    Возвращает (удалено_строк, список_удалённых_файлов).
    """
    deleted_rows = 0
    deleted_files: list[str] = []
    for (year,) in conn.execute(
        'SELECT DISTINCT year FROM tenant_movements ORDER BY year'
    ).fetchall():
        rows = conn.execute(
            'SELECT source_file, COUNT(*) AS n FROM tenant_movements '
            'WHERE year = ? GROUP BY source_file ORDER BY n DESC, source_file',
            (year,),
        ).fetchall()
        if len(rows) <= 1:
            continue
        canonical = rows[0][0]
        for src, n in rows[1:]:
            d = conn.execute(
                'DELETE FROM tenant_movements WHERE year = ? AND source_file = ?',
                (year, src),
            ).rowcount
            deleted_rows += d
            deleted_files.append(src)
            logging.info('  [%d] оставляем %s (%d записей), убираем %s (%d записей)',
                         year, Path(canonical).name, rows[0][1], Path(src).name, n)
    return deleted_rows, deleted_files


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument('--root', default=DEFAULT_ROOT,
                    help=f'Корневая папка с годовыми Excel-файлами (default: {DEFAULT_ROOT})')
    ap.add_argument('--year', type=int, help='Импортировать только один конкретный год')
    ap.add_argument('--dry-run', action='store_true', help='Только сканировать, ничего не писать')
    ap.add_argument('--scan-all', action='store_true',
                    help='Заходить во все .xlsx (а не только в "главные" план-факт файлы)')
    ap.add_argument('--scan-all-sheets', action='store_true',
                    help='Сканировать все листы каждого файла (а не только с "съезд"/"заезд" в имени)')
    ap.add_argument('--no-dedupe', action='store_true',
                    help='Не выкидывать дубликатные источники (оставить все версии файлов)')
    args = ap.parse_args()

    setup_logging()
    logging.info('Импорт съездов/заездов. Корень: %s', args.root)
    logging.info('БД: %s%s', DB_PATH, ' (DRY-RUN)' if args.dry_run else '')

    root = Path(args.root)
    if not root.exists():
        logging.error('Корень не найден или нет доступа: %s', root)
        return 2

    files = discover_files(root, args.year, scan_all=args.scan_all)
    logging.info('Файлов к обработке: %d', len(files))
    if not files:
        return 0

    conn = open_db()
    totals = {'inserted': 0, 'updated': 0, 'skipped': 0, 'sheets_with_data': 0,
              'files_with_data': 0}
    try:
        for path, year in files:
            logging.info('[%d] %s', year, path.name)
            cnt = process_file(conn, path, year, args.dry_run,
                               scan_all_sheets=args.scan_all_sheets)
            for k, v in cnt.items():
                totals[k] = totals.get(k, 0) + v
            if cnt['sheets_with_data'] > 0:
                totals['files_with_data'] += 1
            if not args.dry_run:
                conn.commit()
    finally:
        conn.close()

    logging.info('Готово. Файлов с данными: %d. Записей: вставлено %d, обновлено %d, пропущено %d.',
                 totals['files_with_data'], totals['inserted'], totals['updated'], totals['skipped'])

    if not args.dry_run and not args.no_dedupe:
        logging.info('Дедуп источников по годам:')
        conn = open_db()
        try:
            deleted_rows, deleted_files = dedupe_per_year(conn)
            conn.commit()
            logging.info('Удалено %d строк (дубликатные копии файлов: %d).',
                         deleted_rows, len(deleted_files))
        finally:
            conn.close()
    return 0


if __name__ == '__main__':
    sys.exit(main())
