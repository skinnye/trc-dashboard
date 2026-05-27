"""
Импорт товарооборота арендаторов из листа «НОВАЯ» файла
\\\\Acad-server\\общие\\02_Бухгалтерия\\01_Бюджет\\2026\\02_ТО АП.xlsx.

Назначение: годовая аналитика ТО — рейтинг по обороту, эффективность
по м² (ТО/м²), сводка по категориям, динамика по годам.

Запуск:
  python import_turnover.py
  python import_turnover.py --year 2026
  python import_turnover.py --dry-run
"""
from __future__ import annotations

import argparse
import logging
import os
import re
import sqlite3
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError:
    print('Нужен openpyxl: pip install openpyxl', file=sys.stderr)
    sys.exit(2)

HERE = Path(__file__).parent.resolve()
DB_PATH = Path(os.environ.get('DASHBOARD_DB', str(HERE.parent / 'dashboard.db')))
DEFAULT_FILE = r'\\Acad-server\общие\02_Бухгалтерия\01_Бюджет\2026\02_ТО АП.xlsx'
LOG_DIR = HERE / 'logs'
LOG_DIR.mkdir(exist_ok=True)


# Маппинг колонок листа «НОВАЯ» (1-based). Соответствует первой строке
# с заголовками. Если шапка изменится — обновить здесь.
COLS = {
    'year':              1,   # ГОД
    'arendator':         2,   # Арендатор
    'store_name':        3,   # Название магазина
    'category':          4,   # Категория
    'area_m2':           5,   # S
    'rate_fixed':        6,   # Ставка фикс
    'rate_fixed_to':     7,   # Ставка фикс + ТО
    'ap_fixed':          8,   # АП фикс
    'ap_fixed_indexed':  9,   # АП фикс после индексации
    'ap_extra':         10,   # Доп. АП
    'to_percent':       11,   # % с ТО
    'ap_with_to':       12,   # АП с учётом ТО
    'ap_share_in_to':   13,   # Доля АП в ТО
    'to_sum_year':      14,   # ТО сумма за год/текущая
    'to_sum_period':    15,   # ТО сумма за период янв-ноя
    'to_avg_monthly':   16,   # ТО средний
    'to_per_m2':        17,   # ТО с м2
    'avg_traffic':      18,   # Ср. трафик
    'avg_purchases':    19,   # Ср. кол-во покупок
    'avg_check':        20,   # Ср. чек
    'ap_change_note':   21,   # Изменение условий АП
    'to_yoy_pct':       22,   # Сравнение ср. ТО vs пред. год
}

# Помесячные колонки идут блоками по 6 штук, начиная с колонки 23.
# Структура каждого блока (для месяца M, базовая колонка = MONTHLY_BASE + (M-1)*6):
#   +0: ТО за месяц
#   +1: % к этому месяцу пред. года (yoy)
#   +2: ТО с м² за месяц
#   +3: % к прошлому месяцу (mom)
#   +4: Кол-во покупок за месяц
#   +5: АП за месяц
# Итого 12 × 6 = 72 колонки (23..94).
MONTHLY_BASE = 23
MONTHLY_FIELDS = ('to_sum', 'yoy_pct', 'to_per_m2', 'mom_pct', 'purchases', 'ap_for_month')


def monthly_col(month: int, field_name: str) -> int:
    """Возвращает 1-based индекс колонки для указанного месяца и поля."""
    if not 1 <= month <= 12:
        raise ValueError(f'month must be 1..12, got {month}')
    offset = MONTHLY_FIELDS.index(field_name)
    return MONTHLY_BASE + (month - 1) * 6 + offset


SCHEMA_SQL = '''
CREATE TABLE IF NOT EXISTS turnover_yearly (
  id                 INTEGER PRIMARY KEY AUTOINCREMENT,
  year               INTEGER NOT NULL,
  arendator          TEXT,
  store_name         TEXT NOT NULL,
  category           TEXT,
  area_m2            REAL,
  rate_fixed         REAL,
  rate_fixed_to      REAL,
  ap_fixed           REAL,
  ap_fixed_indexed   REAL,
  ap_extra           REAL,
  to_percent         REAL,
  ap_with_to         REAL,
  ap_share_in_to     REAL,
  to_sum_year        REAL,
  to_sum_period      REAL,
  to_avg_monthly     REAL,
  to_per_m2          REAL,
  avg_traffic        REAL,
  avg_purchases      REAL,
  avg_check          REAL,
  ap_change_note     TEXT,
  to_yoy_pct         REAL,
  source_file        TEXT NOT NULL,
  imported_at        TEXT NOT NULL,
  UNIQUE(year, store_name)
);
CREATE INDEX IF NOT EXISTS idx_turnover_year     ON turnover_yearly(year);
CREATE INDEX IF NOT EXISTS idx_turnover_category ON turnover_yearly(category);
CREATE INDEX IF NOT EXISTS idx_turnover_store    ON turnover_yearly(store_name);

CREATE TABLE IF NOT EXISTS turnover_monthly (
  id              INTEGER PRIMARY KEY AUTOINCREMENT,
  year            INTEGER NOT NULL,
  month           INTEGER NOT NULL,
  store_name      TEXT NOT NULL,
  category        TEXT,
  area_m2         REAL,
  to_sum          REAL,
  to_per_m2       REAL,
  yoy_pct         REAL,
  mom_pct         REAL,
  purchases       REAL,
  ap_for_month    REAL,
  source_file     TEXT NOT NULL,
  imported_at     TEXT NOT NULL,
  UNIQUE(year, month, store_name)
);
CREATE INDEX IF NOT EXISTS idx_turnover_monthly_year_month ON turnover_monthly(year, month);
CREATE INDEX IF NOT EXISTS idx_turnover_monthly_store      ON turnover_monthly(store_name);
'''


def setup_logging() -> Path:
    log_file = LOG_DIR / f'turnover_{datetime.now():%Y-%m-%d_%H-%M}.log'
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s [%(levelname)s] %(message)s',
        handlers=[
            logging.FileHandler(log_file, encoding='utf-8'),
            logging.StreamHandler(sys.stdout),
        ],
        force=True,
    )
    return log_file


def now_iso() -> str:
    return datetime.now().strftime('%Y-%m-%d %H:%M:%S')


def cell_to_float(v: Any) -> float | None:
    """Достаёт число из ячейки openpyxl. Excel formulas с `value=None` и
    cached-результатом в `result` обрабатываются корректно — мы открываем
    workbook с data_only=True, тогда ячейка отдаёт уже посчитанное число."""
    if v is None or v == '':
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v) if v == v else None
    s = str(v).strip().replace(',', '.').replace('\xa0', '').replace(' ', '').rstrip('%')
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def cell_to_int(v: Any) -> int | None:
    f = cell_to_float(v)
    return int(f) if f is not None else None


def cell_to_str(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    # Иногда вместо строки приходит структура {"text": "..."} или формула.
    if s.startswith('{') and 'text' in s:
        m = re.search(r'"text"\s*:\s*"([^"]+)"', s)
        if m:
            s = m.group(1).strip()
    return s or None


def open_db() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.execute('PRAGMA journal_mode = WAL')
    conn.executescript(SCHEMA_SQL)
    return conn


def upsert_yearly(conn: sqlite3.Connection, fields: dict, source_file: str,
                  dry_run: bool) -> str:
    """UPSERT в turnover_yearly. Используем ON CONFLICT по UNIQUE(year, store_name)
    — это атомарно и без TOCTOU-окна (см. code review)."""
    if dry_run:
        return 'inserted'
    now = now_iso()
    cols = list(fields.keys()) + ['source_file', 'imported_at']
    vals = [fields[k] for k in fields] + [source_file, now]
    placeholders = ', '.join(['?'] * len(cols))
    update_cols = [c for c in cols if c not in ('year', 'store_name')]
    update_set = ', '.join(f'{c} = excluded.{c}' for c in update_cols)
    cur = conn.execute(
        f'INSERT INTO turnover_yearly ({", ".join(cols)}) VALUES ({placeholders}) '
        f'ON CONFLICT(year, store_name) DO UPDATE SET {update_set}',
        vals,
    )
    return 'inserted' if cur.rowcount == 1 else 'updated'


def upsert_monthly(conn: sqlite3.Connection, fields: dict, source_file: str,
                   dry_run: bool) -> str:
    if dry_run:
        return 'inserted'
    now = now_iso()
    cols = list(fields.keys()) + ['source_file', 'imported_at']
    vals = [fields[k] for k in fields] + [source_file, now]
    placeholders = ', '.join(['?'] * len(cols))
    update_cols = [c for c in cols if c not in ('year', 'month', 'store_name')]
    update_set = ', '.join(f'{c} = excluded.{c}' for c in update_cols)
    conn.execute(
        f'INSERT INTO turnover_monthly ({", ".join(cols)}) VALUES ({placeholders}) '
        f'ON CONFLICT(year, month, store_name) DO UPDATE SET {update_set}',
        vals,
    )
    return 'inserted'


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument('--file', default=DEFAULT_FILE)
    ap.add_argument('--sheet', default='НОВАЯ')
    ap.add_argument('--year', type=int, help='Только один год')
    ap.add_argument('--dry-run', action='store_true')
    args = ap.parse_args()

    setup_logging()
    logging.info('Импорт товарооборота. Файл: %s', args.file)

    path = Path(args.file)
    if not path.exists():
        logging.error('Файл не найден: %s', path)
        return 2

    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        logging.error('Не удалось открыть: %s', e)
        return 2

    try:
        ws = wb[args.sheet]
    except KeyError:
        logging.error('Лист "%s" не найден. Доступные: %s', args.sheet, wb.sheetnames)
        wb.close()
        return 2

    conn = open_db() if not args.dry_run else None
    totals = {'yearly_in': 0, 'yearly_up': 0, 'monthly_in': 0, 'skipped': 0}

    try:
        # Идём со второй строки — первая это шапка
        for raw_row in ws.iter_rows(min_row=2, values_only=True):
            row = list(raw_row)
            # Безопасный доступ — лист может быть уже шире, чем нам нужно
            def cell(col: int) -> Any:
                return row[col - 1] if col - 1 < len(row) else None

            year = cell_to_int(cell(COLS['year']))
            if year is None or year < 2010 or year > 2040:
                totals['skipped'] += 1
                continue
            if args.year is not None and year != args.year:
                continue
            store_name = cell_to_str(cell(COLS['store_name']))
            if not store_name:
                totals['skipped'] += 1
                continue
            # Нормализуем NBSP и схлопываем повторные пробелы (см. code review).
            store_name = re.sub(r'\s+', ' ', store_name.replace('\xa0', ' ')).strip()
            category = cell_to_str(cell(COLS['category']))
            area_m2 = cell_to_float(cell(COLS['area_m2']))

            yearly_fields = {
                'year': year,
                'arendator':        cell_to_str(cell(COLS['arendator'])),
                'store_name':       store_name,
                'category':         category,
                'area_m2':          area_m2,
                'rate_fixed':       cell_to_float(cell(COLS['rate_fixed'])),
                'rate_fixed_to':    cell_to_float(cell(COLS['rate_fixed_to'])),
                'ap_fixed':         cell_to_float(cell(COLS['ap_fixed'])),
                'ap_fixed_indexed': cell_to_float(cell(COLS['ap_fixed_indexed'])),
                'ap_extra':         cell_to_float(cell(COLS['ap_extra'])),
                'to_percent':       cell_to_float(cell(COLS['to_percent'])),
                'ap_with_to':       cell_to_float(cell(COLS['ap_with_to'])),
                'ap_share_in_to':   cell_to_float(cell(COLS['ap_share_in_to'])),
                'to_sum_year':      cell_to_float(cell(COLS['to_sum_year'])),
                'to_sum_period':    cell_to_float(cell(COLS['to_sum_period'])),
                'to_avg_monthly':   cell_to_float(cell(COLS['to_avg_monthly'])),
                'to_per_m2':        cell_to_float(cell(COLS['to_per_m2'])),
                'avg_traffic':      cell_to_float(cell(COLS['avg_traffic'])),
                'avg_purchases':    cell_to_float(cell(COLS['avg_purchases'])),
                'avg_check':        cell_to_float(cell(COLS['avg_check'])),
                'ap_change_note':   cell_to_str(cell(COLS['ap_change_note'])),
                'to_yoy_pct':       cell_to_float(cell(COLS['to_yoy_pct'])),
            }

            if conn:
                res = upsert_yearly(conn, yearly_fields, str(path), args.dry_run)
                totals['yearly_' + ('in' if res == 'inserted' else 'up')] += 1
            else:
                totals['yearly_in'] += 1

            # ── помесячные данные ──
            # Для каждого месяца проверяем, что есть хоть один из показателей.
            # Если все шесть NULL — месяц не импортируем (магазина в этом
            # месяце ещё/уже не было).
            for m in range(1, 13):
                month_fields = {}
                for f in MONTHLY_FIELDS:
                    month_fields[f] = cell_to_float(cell(monthly_col(m, f)))
                if all(v is None for v in month_fields.values()):
                    continue
                rec = {
                    'year': year, 'month': m, 'store_name': store_name,
                    'category': category, 'area_m2': area_m2,
                    **month_fields,
                }
                if conn:
                    upsert_monthly(conn, rec, str(path), args.dry_run)
                totals['monthly_in'] += 1
    finally:
        wb.close()
        if conn:
            conn.commit()
            conn.close()

    logging.info('Готово. Годовых: вставлено %d, обновлено %d. Помесячных: %d. Пропущено: %d.',
                 totals['yearly_in'], totals['yearly_up'], totals['monthly_in'], totals['skipped'])
    return 0


if __name__ == '__main__':
    sys.exit(main())
